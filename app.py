# -*- coding: utf-8 -*-
"""
SISTEMA GOULARTH DE TORNEIOS — aplicação Flask
=================================================
"""

import os
import sys
import subprocess
import functools
import datetime
import threading
import time
import logging
from flask import (
    Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
)

# Tenta importar psycopg2, se não conseguir, tenta instalar
try:
    import psycopg2
    print("✅ psycopg2 importado com sucesso!")
except ImportError:
    print("⚠️ psycopg2 não encontrado. Tentando instalar...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "psycopg2-binary==2.9.9"])
    import psycopg2
    print("✅ psycopg2 instalado com sucesso!")

import banco
import regras

# Configurar logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "chave-desenvolvimento-trocar")

db = banco.BancoClube(os.environ.get("DB_PATH", "clube.db"))

NOMES_MODALIDADE = {"FIBRA": "Fibra", "CANTO_LIVRE": "Canto Livre"}
NOMES_CATEGORIA = {"FILHOTE": "Filhote", "ADULTO": "Adulto", "MISTO": "Misto"}


# ================================================================
# AUTENTICAÇÃO
# ================================================================
def login_obrigatorio(view):
    @functools.wraps(view)
    def wrapper(*args, **kwargs):
        if "socio_id" not in session:
            flash("Faça login para continuar.", "erro")
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapper


def admin_obrigatorio(view):
    @functools.wraps(view)
    def wrapper(*args, **kwargs):
        if "admin_id" not in session:
            flash("Acesso restrito à administração.", "erro")
            return redirect(url_for("admin_login"))
        return view(*args, **kwargs)
    return wrapper


def admin_com_permissoes(permissoes_requeridas):
    """Decorador que verifica se o admin tem permissão para acessar a rota."""
    def decorator(view):
        @functools.wraps(view)
        def wrapper(*args, **kwargs):
            if "admin_id" not in session:
                flash("Acesso restrito à administração.", "erro")
                return redirect(url_for("admin_login"))
            
            admin_id = session["admin_id"]
            admin = db.obter_admin(admin_id)
            if not admin:
                flash("Administrador não encontrado.", "erro")
                return redirect(url_for("admin_login"))
            
            nivel = admin["nivel"]
            
            if nivel == 1:
                return view(*args, **kwargs)
            
            for permissao in permissoes_requeridas:
                if permissao == "criar_admin" and nivel != 1:
                    flash("Apenas Super Administradores podem criar novos admins.", "erro")
                    return redirect(url_for("admin_dashboard"))
                if permissao == "gerenciar_admins" and nivel > 2:
                    flash("Você não tem permissão para gerenciar administradores.", "erro")
                    return redirect(url_for("admin_dashboard"))
                if permissao == "atualizar_sistema" and nivel != 1:
                    flash("Apenas Super Administradores podem atualizar o sistema.", "erro")
                    return redirect(url_for("admin_dashboard"))
            
            return view(*args, **kwargs)
        return wrapper
    return decorator


@app.context_processor
def injetar_globais():
    return {
        "nomes_modalidade": NOMES_MODALIDADE,
        "nomes_categoria": NOMES_CATEGORIA,
        "ano_atual": datetime.date.today().year,
        "nome_site": "Sistema Goularth de Torneios",
        "subtitulo": "A Evolução do Canário Doméstico"
    }


# ================================================================
# HEALTH CHECK
# ================================================================
@app.route("/health")
def health_check():
    """Endpoint para verificar saúde da aplicação."""
    try:
        db_exists = os.path.exists("clube.db")
        db_size = os.path.getsize("clube.db") if db_exists else 0
        total_resultados = db.contar_resultados_importados() if db_exists else 0
        
        return jsonify({
            "status": "ok",
            "banco_existe": db_exists,
            "banco_tamanho": f"{db_size / 1024:.2f} KB",
            "resultados_importados": total_resultados,
            "timestamp": datetime.datetime.now().isoformat(),
            "usando_postgres": db.usar_postgres if hasattr(db, 'usar_postgres') else False
        })
    except Exception as e:
        logger.error(f"Erro no health check: {e}")
        return jsonify({
            "status": "erro",
            "erro": str(e)
        }), 500


# ================================================================
# PÁGINAS PÚBLICAS
# ================================================================
@app.route("/")
def home():
    torneios = db.listar_torneios()
    festivos = db.listar_festivos()
    return render_template("home.html", torneios=torneios, festivos=festivos)


@app.route("/torneio/<int:torneio_id>")
def ver_torneio(torneio_id):
    torneio = db.obter_torneio(torneio_id)
    if torneio is None:
        flash("Torneio não encontrado.", "erro")
        return redirect(url_for("home"))
    etapas = db.listar_etapas_do_torneio(torneio_id)
    return render_template("torneio_publico.html", torneio=torneio, etapas=etapas)


@app.route("/etapa/<int:etapa_id>")
def ver_etapa_publica(etapa_id):
    etapa = db.obter_etapa(etapa_id)
    if etapa is None:
        flash("Etapa não encontrada.", "erro")
        return redirect(url_for("home"))
    inscritos = db.listar_inscritos_na_etapa(etapa_id)
    return render_template("etapa_publica.html", etapa=etapa, inscritos=inscritos)


# ================================================================
# RESULTADOS E RANKING (PÚBLICOS)
# ================================================================
@app.route("/resultados")
def listar_resultados():
    """Página com todas as etapas importadas."""
    categorias = db.obter_categorias_disponiveis()
    etapas = db.listar_etapas_importadas()
    return render_template("resultados.html", etapas=etapas, categorias=categorias)


@app.route("/resultados/<categoria>")
def listar_resultados_categoria(categoria):
    """Página com resultados filtrados por categoria."""
    etapas = db.listar_etapas_importadas(categoria)
    return render_template("resultados_categoria.html", etapas=etapas, categoria=categoria)


@app.route("/resultados/etapa/<torneio_nome>/<data_etapa>/<categoria>")
def ver_resultado_etapa(torneio_nome, data_etapa, categoria):
    """Página detalhada de uma etapa."""
    resultados = db.obter_resultado_etapa_importado(torneio_nome, data_etapa, categoria)
    return render_template("resultado_etapa.html", 
                          resultados=resultados, 
                          torneio_nome=torneio_nome,
                          data_etapa=data_etapa,
                          categoria=categoria)


@app.route("/ranking")
def ranking_geral():
    """Página com ranking geral do campeonato."""
    try:
        ranking = db.obter_ranking_geral_importado()
        categorias = db.obter_categorias_disponiveis()
        return render_template("ranking_geral.html", ranking=ranking, categorias=categorias)
    except Exception as e:
        logger.error(f"Erro no ranking: {e}")
        return render_template("ranking_geral.html", ranking=[], categorias=[])


@app.route("/ranking/<categoria>")
def ranking_categoria(categoria):
    """Página com ranking por categoria."""
    ranking = db.obter_ranking_geral_importado(categoria)
    categorias = db.obter_categorias_disponiveis()
    return render_template("ranking_categoria.html", 
                          ranking=ranking, 
                          categoria=categoria,
                          categorias=categorias)


# ================================================================
# CRIADORES
# ================================================================
@app.route("/criadores")
def listar_criadores():
    """Página com lista de criadores."""
    busca = request.args.get("busca", "").strip()
    criadores = db.listar_criadores(busca if busca else None)
    return render_template("criadores.html", criadores=criadores, busca=busca)


@app.route("/criador/<int:socio_id>")
def ver_criador(socio_id):
    """Página de detalhes de um criador."""
    criador = db.obter_criador(socio_id)
    if criador is None:
        flash("Criador não encontrado ou não permite exibição pública.", "erro")
        return redirect(url_for("listar_criadores"))
    return render_template("criador_detalhe.html", criador=criador)


# ================================================================
# CADASTRO E LOGIN (SÓCIO)
# ================================================================
@app.route("/cadastro", methods=["GET", "POST"])
def cadastro():
    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        cpf = request.form.get("cpf", "").strip()
        nascimento = request.form.get("nascimento", "")
        sexo = request.form.get("sexo", "")
        criatorio = request.form.get("criatorio", "").strip()
        sigla_clube = request.form.get("sigla_clube", "").strip().upper()
        numero_socio = request.form.get("numero_socio", "").strip()
        cep = request.form.get("cep", "").strip()
        endereco = request.form.get("endereco", "").strip()
        numero = request.form.get("numero", "").strip()
        complemento = request.form.get("complemento", "").strip()
        bairro = request.form.get("bairro", "").strip()
        cidade = request.form.get("cidade", "").strip()
        uf = request.form.get("uf", "").strip().upper()
        pais = request.form.get("pais", "").strip()
        ddi = request.form.get("ddi", "").strip()
        ddd = request.form.get("ddd", "").strip()
        celular = request.form.get("celular", "").strip()
        whatsapp = request.form.get("whatsapp", "").strip()
        email = request.form.get("email", "").strip()
        facebook = request.form.get("facebook", "").strip()
        instagram = request.form.get("instagram", "").strip()
        youtube = request.form.get("youtube", "").strip()
        exibir_dados = request.form.get("exibir_dados", "0") == "1"
        senha = request.form.get("senha", "")
        confirmar = request.form.get("confirmar_senha", "")

        if senha != confirmar:
            flash("As senhas não coincidem.", "erro")
            return render_template("cadastro.html", **request.form)

        try:
            socio_id = db.criar_socio(
                nome=nome, cpf=cpf, nascimento=nascimento, sexo=sexo,
                criatorio=criatorio,
                sigla_clube=sigla_clube, numero_socio=numero_socio,
                cep=cep, endereco=endereco, numero=numero, complemento=complemento,
                bairro=bairro, cidade=cidade, uf=uf, pais=pais,
                ddi=ddi, ddd=ddd, celular=celular, whatsapp=whatsapp,
                email=email, facebook=facebook, instagram=instagram,
                youtube=youtube, exibir_dados=exibir_dados, senha=senha
            )
            session["socio_id"] = socio_id
            session["socio_nome"] = nome
            flash("Cadastro realizado com sucesso! Bem-vindo(a).", "sucesso")
            return redirect(url_for("area_socio"))
        except (regras.ErroValidacao, ValueError) as e:
            flash(str(e), "erro")
            return render_template("cadastro.html", **request.form)

    return render_template("cadastro.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        cpf = request.form.get("cpf", "").strip()
        senha = request.form.get("senha", "")
        
        if not cpf:
            flash("Por favor, digite seu CPF.", "erro")
            return render_template("login.html", cpf=cpf)
        
        try:
            socio = db.autenticar_socio(cpf, senha)
            if socio is None:
                flash("CPF ou senha incorretos.", "erro")
                return render_template("login.html", cpf=cpf)
            
            session["socio_id"] = socio["id"]
            session["socio_nome"] = socio["nome"]
            return redirect(url_for("area_socio"))
            
        except regras.ErroValidacao as e:
            flash(str(e), "erro")
            return render_template("login.html", cpf=cpf)
        except Exception as e:
            logger.error(f"Erro no login: {e}")
            flash("Erro ao fazer login. Tente novamente.", "erro")
            return render_template("login.html", cpf=cpf)
    
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("home"))


# ================================================================
# ADMIN - LOGIN E GESTÃO
# ================================================================
@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        senha = request.form.get("senha", "")
        admin = db.autenticar_admin(email, senha)
        if admin is None:
            flash("Email ou senha incorretos.", "erro")
            return render_template("admin_login.html", email=email)
        session["admin_id"] = admin["id"]
        session["admin_nome"] = admin["nome"]
        session["admin_nivel"] = admin["nivel"]
        return redirect(url_for("admin_dashboard"))
    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_id", None)
    session.pop("admin_nome", None)
    session.pop("admin_nivel", None)
    return redirect(url_for("home"))


@app.route("/admin/trocar-senha", methods=["GET", "POST"])
@admin_obrigatorio
def admin_trocar_senha():
    if request.method == "POST":
        senha_atual = request.form.get("senha_atual", "")
        nova_senha = request.form.get("nova_senha", "")
        confirmar = request.form.get("confirmar_senha", "")
        
        if nova_senha != confirmar:
            flash("As senhas não coincidem.", "erro")
            return render_template("admin_trocar_senha.html")
        
        try:
            db.trocar_senha_admin(session["admin_id"], nova_senha)
            flash("Senha alterada com sucesso!", "sucesso")
            return redirect(url_for("admin_dashboard"))
        except regras.ErroValidacao as e:
            flash(str(e), "erro")
            return render_template("admin_trocar_senha.html")
    
    return render_template("admin_trocar_senha.html")


# ================================================================
# ADMIN - DASHBOARD E TORNEIOS
# ================================================================
@app.route("/admin")
@admin_obrigatorio
def admin_dashboard():
    admin_id = session["admin_id"]
    admin = db.obter_admin(admin_id)
    
    torneios = db.listar_torneios()
    festivos = db.listar_festivos()
    edicoes = db.listar_edicoes_pendentes()
    admins = db.listar_admins() if admin["nivel"] <= 2 else []
    
    return render_template("admin_dashboard.html", 
                          torneios=torneios, 
                          festivos=festivos, 
                          edicoes=edicoes,
                          admins=admins,
                          admin_nivel=admin["nivel"])


@app.route("/admin/criar", methods=["GET", "POST"])
@admin_com_permissoes(["criar_admin"])
def admin_criar():
    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        email = request.form.get("email", "").strip()
        senha = request.form.get("senha", "")
        confirmar = request.form.get("confirmar_senha", "")
        nivel = int(request.form.get("nivel", 2))
        
        if senha != confirmar:
            flash("As senhas não coincidem.", "erro")
            return render_template("admin_criar.html", nome=nome, email=email)
        
        try:
            db.criar_admin(nome, email, senha, nivel)
            flash(f"Administrador '{nome}' criado com sucesso! Nível: {nivel}", "sucesso")
            return redirect(url_for("admin_dashboard"))
        except (regras.ErroValidacao, ValueError) as e:
            flash(str(e), "erro")
            return render_template("admin_criar.html", nome=nome, email=email)
    
    return render_template("admin_criar.html")


@app.route("/admin/gerenciar")
@admin_com_permissoes(["gerenciar_admins"])
def admin_gerenciar():
    admins = db.listar_admins()
    return render_template("admin_gerenciar.html", admins=admins)


@app.route("/admin/excluir/<int:admin_id>", methods=["POST"])
@admin_com_permissoes(["criar_admin"])
def admin_excluir(admin_id):
    try:
        db.excluir_admin(admin_id)
        flash("Administrador excluído com sucesso.", "sucesso")
    except ValueError as e:
        flash(str(e), "erro")
    return redirect(url_for("admin_gerenciar"))


@app.route("/admin/alterar-nivel/<int:admin_id>", methods=["POST"])
@admin_com_permissoes(["criar_admin"])
def admin_alterar_nivel(admin_id):
    try:
        novo_nivel = int(request.form.get("nivel", 2))
        db.alterar_nivel_admin(admin_id, novo_nivel, session["admin_id"])
        flash(f"Nível alterado para {novo_nivel}.", "sucesso")
    except ValueError as e:
        flash(str(e), "erro")
    return redirect(url_for("admin_gerenciar"))


@app.route("/admin/atualizar-sistema")
@admin_com_permissoes(["atualizar_sistema"])
def admin_atualizar_sistema():
    flash("Sistema atualizado com sucesso!", "sucesso")
    return redirect(url_for("admin_dashboard"))


# ================================================================
# ADMIN - TORNEIOS
# ================================================================
@app.route("/admin/torneios/novo", methods=["GET", "POST"])
@admin_obrigatorio
def admin_novo_torneio():
    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        endereco = request.form.get("endereco", "").strip()
        mapa_url = request.form.get("mapa_url", "").strip()
        try:
            torneio_id = db.criar_torneio(nome, endereco, mapa_url)
            flash(f"Torneio '{nome}' criado com sucesso!", "sucesso")
            return redirect(url_for("admin_ver_torneio", torneio_id=torneio_id))
        except regras.ErroValidacao as e:
            flash(str(e), "erro")
    return render_template("admin_novo_torneio.html")


@app.route("/admin/torneios/<int:torneio_id>/editar", methods=["GET", "POST"])
@admin_obrigatorio
def admin_editar_torneio(torneio_id):
    torneio = db.obter_torneio(torneio_id)
    if torneio is None:
        flash("Torneio não encontrado.", "erro")
        return redirect(url_for("admin_dashboard"))
    
    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        endereco = request.form.get("endereco", "").strip()
        mapa_url = request.form.get("mapa_url", "").strip()
        try:
            db.editar_torneio(torneio_id, nome, endereco, mapa_url)
            flash("Torneio atualizado com sucesso!", "sucesso")
            return redirect(url_for("admin_ver_torneio", torneio_id=torneio_id))
        except regras.ErroValidacao as e:
            flash(str(e), "erro")
    
    return render_template("admin_editar_torneio.html", torneio=torneio)


@app.route("/admin/torneios/<int:torneio_id>")
@admin_obrigatorio
def admin_ver_torneio(torneio_id):
    torneio = db.obter_torneio(torneio_id)
    if torneio is None:
        flash("Torneio não encontrado.", "erro")
        return redirect(url_for("admin_dashboard"))
    etapas = db.listar_etapas_do_torneio(torneio_id)
    return render_template("admin_ver_torneio.html", torneio=torneio, etapas=etapas)


@app.route("/admin/torneios/<int:torneio_id>/etapa/nova", methods=["GET", "POST"])
@admin_obrigatorio
def admin_nova_etapa(torneio_id):
    torneio = db.obter_torneio(torneio_id)
    if torneio is None:
        flash("Torneio não encontrado.", "erro")
        return redirect(url_for("admin_dashboard"))

    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        modalidade = request.form.get("modalidade", "").strip().upper()
        categoria = request.form.get("categoria", "").strip().upper()
        data_etapa = request.form.get("data_etapa", "")
        prazo_inscricao = request.form.get("prazo_inscricao", "")
        limite_raw = request.form.get("limite_por_cpf", "").strip()
        limite_por_cpf = int(limite_raw) if limite_raw else None
        limite_inscricoes_raw = request.form.get("limite_inscricoes", "").strip()
        limite_inscricoes = int(limite_inscricoes_raw) if limite_inscricoes_raw else None
        info_pagamento = request.form.get("info_pagamento", "").strip()
        prazo_pagamento = request.form.get("prazo_pagamento", "")

        try:
            db.criar_etapa(
                torneio_id=torneio_id,
                nome=nome,
                modalidade=modalidade,
                categoria=categoria,
                data_etapa=data_etapa,
                prazo_inscricao=prazo_inscricao,
                limite_por_cpf=limite_por_cpf,
                limite_inscricoes=limite_inscricoes,
                info_pagamento=info_pagamento,
                prazo_pagamento=prazo_pagamento
            )
            flash(f"Etapa '{nome}' criada com sucesso!", "sucesso")
            return redirect(url_for("admin_ver_torneio", torneio_id=torneio_id))
        except regras.ErroValidacao as e:
            flash(str(e), "erro")

    return render_template("admin_nova_etapa.html", torneio=torneio)


@app.route("/admin/etapa/<int:etapa_id>/editar", methods=["GET", "POST"])
@admin_obrigatorio
def admin_editar_etapa(etapa_id):
    etapa = db.obter_etapa(etapa_id)
    if etapa is None:
        flash("Etapa não encontrada.", "erro")
        return redirect(url_for("admin_dashboard"))
    
    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        modalidade = request.form.get("modalidade", "").strip().upper()
        categoria = request.form.get("categoria", "").strip().upper()
        data_etapa = request.form.get("data_etapa", "")
        prazo_inscricao = request.form.get("prazo_inscricao", "")
        limite_raw = request.form.get("limite_por_cpf", "").strip()
        limite_por_cpf = int(limite_raw) if limite_raw else None
        limite_inscricoes_raw = request.form.get("limite_inscricoes", "").strip()
        limite_inscricoes = int(limite_inscricoes_raw) if limite_inscricoes_raw else None
        info_pagamento = request.form.get("info_pagamento", "").strip()
        prazo_pagamento = request.form.get("prazo_pagamento", "")
        
        try:
            db.editar_etapa(etapa_id, nome, modalidade, categoria, data_etapa, 
                           prazo_inscricao, limite_por_cpf, limite_inscricoes,
                           info_pagamento, prazo_pagamento)
            flash("Etapa atualizada com sucesso!", "sucesso")
            return redirect(url_for("admin_ver_etapa", etapa_id=etapa_id))
        except regras.ErroValidacao as e:
            flash(str(e), "erro")
    
    return render_template("admin_editar_etapa.html", etapa=etapa)


@app.route("/admin/festivo/novo", methods=["GET", "POST"])
@admin_obrigatorio
def admin_novo_festivo():
    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        modalidade = request.form.get("modalidade", "").strip().upper()
        categoria = request.form.get("categoria", "").strip().upper()
        data_etapa = request.form.get("data_etapa", "")
        prazo_inscricao = request.form.get("prazo_inscricao", "")
        limite_raw = request.form.get("limite_por_cpf", "").strip()
        limite_por_cpf = int(limite_raw) if limite_raw else None
        limite_inscricoes_raw = request.form.get("limite_inscricoes", "").strip()
        limite_inscricoes = int(limite_inscricoes_raw) if limite_inscricoes_raw else None
        info_pagamento = request.form.get("info_pagamento", "").strip()
        prazo_pagamento = request.form.get("prazo_pagamento", "")

        try:
            db.criar_festivo(
                nome=nome,
                modalidade=modalidade,
                categoria=categoria,
                data_etapa=data_etapa,
                prazo_inscricao=prazo_inscricao,
                limite_por_cpf=limite_por_cpf,
                limite_inscricoes=limite_inscricoes,
                info_pagamento=info_pagamento,
                prazo_pagamento=prazo_pagamento
            )
            flash(f"Festivo '{nome}' criado com sucesso!", "sucesso")
            return redirect(url_for("admin_dashboard"))
        except regras.ErroValidacao as e:
            flash(str(e), "erro")

    return render_template("admin_novo_festivo.html")


@app.route("/admin/etapa/<int:etapa_id>")
@admin_obrigatorio
def admin_ver_etapa(etapa_id):
    etapa = db.obter_etapa(etapa_id)
    if etapa is None:
        flash("Etapa não encontrada.", "erro")
        return redirect(url_for("admin_dashboard"))
    
    inscritos = db.listar_inscritos_na_etapa(etapa_id)
    
    # Ordena por ordem (crescente) - quem tem ordem NULL vai para o final
    inscritos_ordenados = sorted(
        inscritos,
        key=lambda x: (x["ordem"] is None, x["ordem"] if x["ordem"] is not None else float('inf'))
    )
    
    total_inscritos = db.contar_inscricoes_na_etapa(etapa_id)
    return render_template("admin_ver_etapa.html", etapa=etapa, inscritos=inscritos_ordenados, total_inscritos=total_inscritos)


@app.route("/admin/etapa/<int:etapa_id>/limite", methods=["POST"])
@admin_obrigatorio
def admin_atualizar_limite(etapa_id):
    etapa = db.obter_etapa(etapa_id)
    if etapa is None:
        flash("Etapa não encontrada.", "erro")
        return redirect(url_for("admin_dashboard"))
    
    try:
        novo_limite = request.form.get("limite_inscricoes", "").strip()
        if novo_limite == "":
            novo_limite = None
        else:
            novo_limite = int(novo_limite)
            if novo_limite < 0:
                raise ValueError("O limite deve ser maior ou igual a 0.")
        
        db.atualizar_limite_inscricoes_etapa(etapa_id, novo_limite)
        if novo_limite is None:
            flash("Limite de inscrições removido (ilimitado).", "sucesso")
        else:
            flash(f"Limite de inscrições atualizado para {novo_limite}.", "sucesso")
    except ValueError as e:
        flash(str(e), "erro")
    
    return redirect(url_for("admin_ver_etapa", etapa_id=etapa_id))


@app.route("/admin/etapa/<int:etapa_id>/toggle_inscricoes", methods=["POST"])
@admin_obrigatorio
def admin_toggle_inscricoes(etapa_id):
    etapa = db.obter_etapa(etapa_id)
    if etapa is None:
        flash("Etapa não encontrada.", "erro")
        return redirect(url_for("admin_dashboard"))
    
    nova_abertura = not etapa["inscricoes_abertas"]
    db.atualizar_abertura_inscricoes(etapa_id, nova_abertura)
    status = "abertas" if nova_abertura else "fechadas"
    flash(f"Inscrições {status} com sucesso.", "sucesso")
    return redirect(url_for("admin_ver_etapa", etapa_id=etapa_id))


# ================================================================
# ADMIN - LISTA DE APRESENTAÇÃO
# ================================================================
@app.route("/admin/etapa/<int:etapa_id>/lista-apresentacao")
@admin_obrigatorio
def admin_lista_apresentacao(etapa_id):
    etapa = db.obter_etapa(etapa_id)
    if etapa is None:
        flash("Etapa não encontrada.", "erro")
        return redirect(url_for("admin_dashboard"))
    
    inscritos = db.listar_inscritos_na_etapa(etapa_id)
    
    # Ordena por ordem (crescente) - quem tem ordem NULL vai para o final
    inscritos_ordenados = sorted(
        inscritos,
        key=lambda x: (x["ordem"] is None, x["ordem"] if x["ordem"] is not None else float('inf'))
    )
    
    lista = []
    for i, ins in enumerate(inscritos_ordenados, 1):
        lista.append({
            "ordem_sequencial": i,
            "ordem_sistema": ins["ordem"] if ins["ordem"] else "—",
            "passaro_nome": ins["passaro_nome"],
            "codigo_ave": ins["codigo_ave"],
            "socio_nome": ins["socio_nome"]
        })
    
    return render_template("admin_lista_apresentacao.html", 
                          etapa=etapa, 
                          lista=lista,
                          modalidade=etapa["modalidade"])


@app.route("/admin/etapa/<int:etapa_id>/exportar-lista")
@admin_obrigatorio
def admin_exportar_lista_apresentacao(etapa_id):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        
        etapa = db.obter_etapa(etapa_id)
        if etapa is None:
            flash("Etapa não encontrada.", "erro")
            return redirect(url_for("admin_dashboard"))
        
        inscritos = db.listar_inscritos_na_etapa(etapa_id)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lista Apresentacao"
        
        headers = ["#", "ORDEM/ESTACA", "PÁSSARO", "ANILHA", "PROPRIETÁRIO"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        for row, ins in enumerate(inscritos, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=ins["ordem"] if ins["ordem"] else "")
            ws.cell(row=row, column=3, value=ins["passaro_nome"])
            ws.cell(row=row, column=4, value=ins["codigo_ave"])
            ws.cell(row=row, column=5, value=ins["socio_nome"])
        
        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        nome_arquivo = f"lista_apresentacao_{etapa['nome']}_{etapa['data_etapa']}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except ImportError:
        flash("Biblioteca 'openpyxl' não instalada. Use: pip install openpyxl", "erro")
        return redirect(url_for("admin_ver_etapa", etapa_id=etapa_id))
    except Exception as e:
        flash(f"Erro ao exportar: {str(e)}", "erro")
        return redirect(url_for("admin_ver_etapa", etapa_id=etapa_id))


@app.route("/admin/etapa/<int:etapa_id>/exportar-lista-livre")
@admin_obrigatorio
def admin_exportar_lista_livre(etapa_id):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        
        etapa = db.obter_etapa(etapa_id)
        if etapa is None:
            flash("Etapa não encontrada.", "erro")
            return redirect(url_for("admin_dashboard"))
        
        if etapa["modalidade"] != "CANTO_LIVRE":
            flash("Esta função é apenas para Canto Livre.", "erro")
            return redirect(url_for("admin_ver_etapa", etapa_id=etapa_id))
        
        inscritos = db.listar_inscritos_na_etapa(etapa_id)
        
        inscritos_ordenados = sorted(
            [ins for ins in inscritos if ins["ordem"] is not None],
            key=lambda x: x["ordem"]
        )
        sem_ordem = [ins for ins in inscritos if ins["ordem"] is None]
        inscritos_ordenados.extend(sem_ordem)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lista Apresentacao"
        
        headers = ["NOME", "ANILHA", "PROPRIETÁRIO"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center")
        
        for row, ins in enumerate(inscritos_ordenados, 2):
            ws.cell(row=row, column=1, value=ins["passaro_nome"])
            ws.cell(row=row, column=2, value=ins["codigo_ave"])
            ws.cell(row=row, column=3, value=ins["socio_nome"])
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        nome_arquivo = f"lista_livre_{etapa['nome']}_{etapa['data_etapa']}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except ImportError:
        flash("Biblioteca 'openpyxl' não instalada. Use: pip install openpyxl", "erro")
        return redirect(url_for("admin_ver_etapa", etapa_id=etapa_id))
    except Exception as e:
        flash(f"Erro ao exportar: {str(e)}", "erro")
        return redirect(url_for("admin_ver_etapa", etapa_id=etapa_id))


# ================================================================
# ADMIN - RESULTADO MANUAL
# ================================================================
@app.route("/admin/etapa/<int:etapa_id>/resultado-manual", methods=["GET", "POST"])
@admin_obrigatorio
def admin_resultado_manual(etapa_id):
    etapa = db.obter_etapa(etapa_id)
    if etapa is None:
        flash("Etapa não encontrada.", "erro")
        return redirect(url_for("admin_dashboard"))
    
    inscritos = db.listar_inscritos_na_etapa(etapa_id)
    resultados = db.obter_resultados_etapa(etapa_id)
    
    resultados_dict = {r["inscricao_id"]: r for r in resultados}
    
    dados = []
    for ins in inscritos:
        dados.append({
            "inscricao_id": ins["inscricao_id"],
            "ordem": ins["ordem"],
            "passaro_nome": ins["passaro_nome"],
            "codigo_ave": ins["codigo_ave"],
            "socio_nome": ins["socio_nome"],
            "resultado": resultados_dict.get(ins["inscricao_id"]),
            "modalidade": etapa["modalidade"]
        })
    
    return render_template("admin_resultado_manual.html", 
                          etapa=etapa, 
                          dados=dados)


@app.route("/admin/resultado-manual/salvar", methods=["POST"])
@admin_obrigatorio
def admin_resultado_manual_salvar():
    try:
        inscricao_id = request.form.get("inscricao_id", type=int)
        posicao = request.form.get("posicao", type=int)
        tempo_parcial = request.form.get("tempo_parcial", "").strip()
        tempo_final = request.form.get("tempo_final", "").strip()
        
        if not inscricao_id or not posicao:
            flash("Posição é obrigatória.", "erro")
            return redirect(request.referrer)
        
        with db._conexao() as conn:
            if db.usar_postgres:
                cur = conn.cursor()
                cur.execute("""
                    SELECT i.*, e.modalidade 
                    FROM inscricoes i
                    JOIN etapas e ON e.id = i.etapa_id
                    WHERE i.id = %s
                """, (inscricao_id,))
                insc = cur.fetchone()
            else:
                insc = conn.execute("""
                    SELECT i.*, e.modalidade 
                    FROM inscricoes i
                    JOIN etapas e ON e.id = i.etapa_id
                    WHERE i.id = ?
                """, (inscricao_id,)).fetchone()
            
            if not insc:
                flash("Inscrição não encontrada.", "erro")
                return redirect(request.referrer)
            
            if insc["modalidade"] == "CANTO_LIVRE":
                if not tempo_final:
                    flash("Tempo Final é obrigatório para Canto Livre.", "erro")
                    return redirect(request.referrer)
                notas = [0, 0, 0, 0, 0]
            else:
                if not tempo_parcial or not tempo_final:
                    flash("Tempo Parcial e Tempo Final são obrigatórios para Fibra.", "erro")
                    return redirect(request.referrer)
                try:
                    notas = [float(tempo_parcial.replace(":", ".")), float(tempo_final.replace(":", ".")), 0, 0, 0]
                except:
                    notas = [0, 0, 0, 0, 0]
        
        db.salvar_resultado_inscricao(inscricao_id, notas, posicao)
        flash("Resultado salvo com sucesso!", "sucesso")
        
    except Exception as e:
        flash(f"Erro ao salvar: {str(e)}", "erro")
    
    return redirect(request.referrer)


@app.route("/admin/resultado-manual/<int:resultado_id>/excluir", methods=["POST"])
@admin_obrigatorio
def admin_resultado_manual_excluir(resultado_id):
    try:
        with db._conexao() as conn:
            if db.usar_postgres:
                cur = conn.cursor()
                cur.execute("DELETE FROM resultados WHERE id = %s", (resultado_id,))
            else:
                conn.execute("DELETE FROM resultados WHERE id = ?", (resultado_id,))
        flash("Resultado excluído com sucesso!", "sucesso")
    except Exception as e:
        flash(f"Erro ao excluir: {str(e)}", "erro")
    return redirect(request.referrer)


# ================================================================
# ADMIN - RESULTADOS
# ================================================================
@app.route("/admin/etapa/<int:etapa_id>/resultados", methods=["GET", "POST"])
@admin_obrigatorio
def admin_resultados_etapa(etapa_id):
    etapa = db.obter_etapa(etapa_id)
    if etapa is None:
        flash("Etapa não encontrada.", "erro")
        return redirect(url_for("admin_dashboard"))
    
    if request.method == "POST":
        for ins in db.listar_inscritos_na_etapa(etapa_id):
            inscricao_id = ins["inscricao_id"]
            classificacao = request.form.get(f"classificacao_{inscricao_id}", type=int)
            
            tempo_parcial = request.form.get(f"tempo_parcial_{inscricao_id}", "").strip()
            tempo_final = request.form.get(f"tempo_final_{inscricao_id}", "").strip()
            
            if classificacao:
                try:
                    if etapa["modalidade"] == "CANTO_LIVRE":
                        notas = [float(tempo_final.replace(":", ".")) if tempo_final else 0, 0, 0, 0, 0]
                    else:
                        notas = [
                            float(tempo_parcial.replace(":", ".")) if tempo_parcial else 0,
                            float(tempo_final.replace(":", ".")) if tempo_final else 0,
                            0, 0, 0
                        ]
                    db.salvar_resultado_inscricao(inscricao_id, notas, classificacao)
                except Exception as e:
                    flash(f"Erro ao salvar {ins['passaro_nome']}: {str(e)}", "erro")
        
        flash("Resultados salvos com sucesso!", "sucesso")
        return redirect(url_for("admin_resultados_etapa", etapa_id=etapa_id))
    
    inscritos = db.listar_inscritos_na_etapa(etapa_id)
    resultados = db.obter_resultados_etapa(etapa_id)
    
    return render_template("admin_resultados_etapa.html", 
                          etapa=etapa, 
                          inscritos=inscritos,
                          resultados=resultados)


@app.route("/admin/resultados")
@admin_obrigatorio
def admin_resultados():
    etapas = db.listar_etapas_importadas()
    return render_template("admin_resultados.html", etapas=etapas)


@app.route("/admin/resultados/importar", methods=["GET", "POST"])
@admin_obrigatorio
def admin_importar_resultado():
    if request.method == "POST":
        try:
            torneio_nome = request.form.get("torneio_nome", "").strip()
            categoria = request.form.get("categoria", "").strip()
            data_etapa = request.form.get("data_etapa", "")
            dados = request.form.get("dados", "").strip()
            
            if not torneio_nome or not categoria or not data_etapa or not dados:
                flash("Todos os campos são obrigatórios.", "erro")
                return render_template("admin_importar_resultado.html")
            
            import re
            linhas = dados.strip().split('\n')
            resultados = []
            
            for linha in linhas:
                if not linha.strip():
                    continue
                partes = [p.strip() for p in linha.split('|')]
                if len(partes) < 6:
                    continue
                
                pos_str = re.sub(r'[°º]', '', partes[0].strip())
                try:
                    posicao = int(pos_str)
                except:
                    continue
                
                passaro_nome = partes[1].strip()
                anilha = partes[2].strip()
                proprietario = partes[3].strip()
                tempo = partes[4].strip()
                
                pontos_str = partes[5].strip()
                try:
                    pontos = int(pontos_str)
                except:
                    pontos = 0
                
                resultados.append({
                    "posicao": posicao,
                    "passaro_nome": passaro_nome,
                    "anilha": anilha,
                    "proprietario": proprietario,
                    "tempo": tempo,
                    "pontos": pontos
                })
            
            if not resultados:
                flash("Nenhum dado válido encontrado. Verifique o formato.", "erro")
                return render_template("admin_importar_resultado.html")
            
            db.importar_resultado_etapa(torneio_nome, categoria, data_etapa, resultados)
            flash(f"Resultado da etapa '{torneio_nome}' importado com sucesso! ({len(resultados)} pássaros)", "sucesso")
            return redirect(url_for("admin_resultados"))
            
        except Exception as e:
            flash(f"Erro ao importar: {str(e)}", "erro")
            return render_template("admin_importar_resultado.html")
    
    return render_template("admin_importar_resultado.html")


@app.route("/admin/resultados/upload", methods=["GET", "POST"])
@admin_obrigatorio
def admin_upload_resultado():
    if request.method == "POST":
        try:
            if 'arquivo' not in request.files:
                flash("Nenhum arquivo selecionado.", "erro")
                return redirect(url_for("admin_upload_resultado"))
            
            arquivo = request.files['arquivo']
            if arquivo.filename == '':
                flash("Nenhum arquivo selecionado.", "erro")
                return redirect(url_for("admin_upload_resultado"))
            
            if not arquivo.filename.lower().endswith(('.csv', '.xlsx')):
                flash("Formato não suportado. Use CSV ou Excel.", "erro")
                return redirect(url_for("admin_upload_resultado"))
            
            import tempfile
            import os
            
            with tempfile.NamedTemporaryFile(delete=False, suffix='.csv' if arquivo.filename.lower().endswith('.csv') else '.xlsx') as tmp:
                arquivo.save(tmp.name)
                caminho_tmp = tmp.name
            
            torneio_nome = request.form.get("torneio_nome", "").strip()
            categoria = request.form.get("categoria", "").strip()
            data_etapa = request.form.get("data_etapa", "")
            
            if not torneio_nome or not categoria or not data_etapa:
                flash("Todos os campos são obrigatórios.", "erro")
                os.unlink(caminho_tmp)
                return render_template("admin_upload_resultado.html")
            
            try:
                if arquivo.filename.lower().endswith('.csv'):
                    db.importar_csv_resultado(caminho_tmp, torneio_nome, categoria, data_etapa)
                else:
                    # Usa openpyxl em vez de pandas
                    from openpyxl import load_workbook
                    wb = load_workbook(caminho_tmp, data_only=True)
                    ws = wb.active
                    
                    resultados = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and row[0] is not None:
                            try:
                                resultados.append({
                                    "posicao": int(row[0]) if row[0] is not None else 0,
                                    "passaro_nome": str(row[2]) if row[2] is not None else "",
                                    "anilha": str(row[3]) if row[3] is not None else "",
                                    "proprietario": str(row[1]) if row[1] is not None else "",
                                    "tempo": str(row[4]) if row[4] is not None else "",
                                    "pontos": int(row[5]) if row[5] is not None else 0
                                })
                            except (ValueError, TypeError):
                                continue
                    
                    if not resultados:
                        raise ValueError("Nenhum dado válido encontrado no Excel.")
                    
                    db.importar_resultado_etapa(torneio_nome, categoria, data_etapa, resultados)
                
                os.unlink(caminho_tmp)
                flash(f"Resultado importado com sucesso!", "sucesso")
                return redirect(url_for("admin_resultados"))
                
            except Exception as e:
                os.unlink(caminho_tmp)
                flash(f"Erro ao importar: {str(e)}", "erro")
                return render_template("admin_upload_resultado.html")
            
        except Exception as e:
            flash(f"Erro: {str(e)}", "erro")
            return render_template("admin_upload_resultado.html")
    
    return render_template("admin_upload_resultado.html")


@app.route("/admin/resultados/editar/<torneio_nome>/<data_etapa>/<categoria>", methods=["GET", "POST"])
@admin_obrigatorio
def admin_editar_resultado(torneio_nome, data_etapa, categoria):
    if request.method == "POST":
        try:
            dados = request.form.get("dados", "").strip()
            
            if not dados:
                flash("Os dados são obrigatórios.", "erro")
                return render_template("admin_editar_resultado.html", 
                                      torneio_nome=torneio_nome, 
                                      data_etapa=data_etapa, 
                                      categoria=categoria)
            
            import re
            linhas = dados.strip().split('\n')
            resultados = []
            
            for linha in linhas:
                if not linha.strip():
                    continue
                partes = [p.strip() for p in linha.split('|')]
                if len(partes) < 6:
                    continue
                
                pos_str = re.sub(r'[°º]', '', partes[0].strip())
                try:
                    posicao = int(pos_str)
                except:
                    continue
                
                passaro_nome = partes[1].strip()
                anilha = partes[2].strip()
                proprietario = partes[3].strip()
                tempo = partes[4].strip()
                
                pontos_str = partes[5].strip()
                try:
                    pontos = int(pontos_str)
                except:
                    pontos = 0
                
                resultados.append({
                    "posicao": posicao,
                    "passaro_nome": passaro_nome,
                    "anilha": anilha,
                    "proprietario": proprietario,
                    "tempo": tempo,
                    "pontos": pontos
                })
            
            if not resultados:
                flash("Nenhum dado válido encontrado.", "erro")
                return render_template("admin_editar_resultado.html", 
                                      torneio_nome=torneio_nome, 
                                      data_etapa=data_etapa, 
                                      categoria=categoria)
            
            db.editar_resultado_etapa(torneio_nome, data_etapa, categoria, resultados)
            flash(f"Resultado da etapa '{torneio_nome}' atualizado com sucesso!", "sucesso")
            return redirect(url_for("admin_resultados"))
            
        except Exception as e:
            flash(f"Erro ao editar: {str(e)}", "erro")
            return render_template("admin_editar_resultado.html", 
                                  torneio_nome=torneio_nome, 
                                  data_etapa=data_etapa, 
                                  categoria=categoria)
    
    resultados = db.obter_resultado_etapa_importado(torneio_nome, data_etapa, categoria)
    return render_template("admin_editar_resultado.html", 
                          torneio_nome=torneio_nome, 
                          data_etapa=data_etapa, 
                          categoria=categoria,
                          resultados=resultados)


@app.route("/admin/resultados/excluir", methods=["POST"])
@admin_obrigatorio
def admin_excluir_etapa():
    torneio_nome = request.form.get("torneio_nome", "").strip()
    data_etapa = request.form.get("data_etapa", "")
    categoria = request.form.get("categoria", "").strip()
    
    try:
        db.excluir_etapa_importada(torneio_nome, data_etapa, categoria)
        flash(f"Etapa '{torneio_nome}' excluída com sucesso!", "sucesso")
    except Exception as e:
        flash(f"Erro ao excluir: {str(e)}", "erro")
    
    return redirect(url_for("admin_resultados"))


@app.route("/admin/resultados/exportar/<torneio_nome>/<data_etapa>/<categoria>")
@admin_obrigatorio
def admin_exportar_inscritos(torneio_nome, data_etapa, categoria):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        
        inscritos = db.exportar_inscritos_para_excel(torneio_nome, data_etapa, categoria)
        
        if not inscritos:
            flash("Nenhum inscrito encontrado.", "erro")
            return redirect(url_for("admin_resultados"))
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inscritos"
        
        headers = ["NOME", "ANILHA", "PROPRIETÁRIO"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        for row, ins in enumerate(inscritos, 2):
            ws.cell(row=row, column=1, value=ins["passaro_nome"])
            ws.cell(row=row, column=2, value=ins["anilha"])
            ws.cell(row=row, column=3, value=ins["proprietario"])
        
        for col in range(1, 4):
            ws.column_dimensions[chr(64 + col)].width = 30
        
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        nome_arquivo = f"inscritos_{categoria}_{data_etapa}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except ImportError:
        flash("Biblioteca 'openpyxl' não instalada. Use: pip install openpyxl", "erro")
        return redirect(url_for("admin_resultados"))
    except Exception as e:
        flash(f"Erro ao exportar: {str(e)}", "erro")
        return redirect(url_for("admin_resultados"))


# ================================================================
# ADMIN - TRANSFERÊNCIAS
# ================================================================
@app.route("/admin/transferencias")
@admin_obrigatorio
def admin_transferencias():
    transferencias = db.listar_transferencias_pendentes()
    return render_template("admin_transferencias.html", transferencias=transferencias)


@app.route("/admin/transferencia/<int:transferencia_id>/aprovar", methods=["POST"])
@admin_obrigatorio
def admin_aprovar_transferencia(transferencia_id):
    try:
        db.aprovar_transferencia_admin(transferencia_id)
        flash("Transferência aprovada com sucesso!", "sucesso")
    except (regras.ErroValidacao, ValueError) as e:
        flash(str(e), "erro")
    return redirect(url_for("admin_transferencias"))


@app.route("/admin/transferencia/<int:transferencia_id>/recusar", methods=["POST"])
@admin_obrigatorio
def admin_recusar_transferencia(transferencia_id):
    try:
        db.recusar_transferencia_admin(transferencia_id)
        flash("Transferência recusada.", "sucesso")
    except (regras.ErroValidacao, ValueError) as e:
        flash(str(e), "erro")
    return redirect(url_for("admin_transferencias"))


# ================================================================
# ADMIN - PAGAMENTOS
# ================================================================
@app.route("/admin/pagamento/<int:inscricao_id>/confirmar", methods=["POST"])
@admin_obrigatorio
def admin_confirmar_pagamento(inscricao_id):
    try:
        db.confirmar_pagamento(inscricao_id)
        flash("Pagamento confirmado com sucesso!", "sucesso")
    except ValueError as e:
        flash(str(e), "erro")
    return redirect(request.referrer or url_for("admin_dashboard"))


@app.route("/admin/pagamento/<int:inscricao_id>/recusar", methods=["POST"])
@admin_obrigatorio
def admin_recusar_pagamento(inscricao_id):
    try:
        db.recusar_pagamento(inscricao_id)
        flash("Pagamento recusado.", "sucesso")
    except ValueError as e:
        flash(str(e), "erro")
    return redirect(request.referrer or url_for("admin_dashboard"))


# ================================================================
# ADMIN - APROVAÇÃO DE EDIÇÕES DE PÁSSAROS
# ================================================================
@app.route("/admin/edicoes/<int:edicao_id>/aprovar", methods=["POST"])
@admin_obrigatorio
def admin_aprovar_edicao(edicao_id):
    try:
        db.aprovar_edicao_passaro(edicao_id)
        flash("Edição aprovada com sucesso!", "sucesso")
    except ValueError as e:
        flash(str(e), "erro")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/edicoes/<int:edicao_id>/rejeitar", methods=["POST"])
@admin_obrigatorio
def admin_rejeitar_edicao(edicao_id):
    try:
        db.rejeitar_edicao_passaro(edicao_id)
        flash("Edição rejeitada.", "sucesso")
    except ValueError as e:
        flash(str(e), "erro")
    return redirect(url_for("admin_dashboard"))


# ================================================================
# ADMIN - CANCELAR INSCRIÇÃO
# ================================================================
@app.route("/admin/inscricao/<int:inscricao_id>/cancelar", methods=["POST"])
@admin_obrigatorio
def admin_cancelar_inscricao(inscricao_id):
    try:
        with db._conexao() as conn:
            if db.usar_postgres:
                cur = conn.cursor()
                cur.execute("SELECT passaro_id FROM inscricoes WHERE id = %s", (inscricao_id,))
                insc = cur.fetchone()
            else:
                insc = conn.execute(
                    "SELECT passaro_id FROM inscricoes WHERE id = ?",
                    (inscricao_id,)
                ).fetchone()
            
            if not insc:
                flash("Inscrição não encontrada.", "erro")
                return redirect(request.referrer or url_for("area_socio"))
            
            if db.usar_postgres:
                cur = conn.cursor()
                cur.execute("SELECT socio_id FROM passaros WHERE id = %s", (insc["passaro_id"],))
                passaro = cur.fetchone()
            else:
                passaro = conn.execute(
                    "SELECT socio_id FROM passaros WHERE id = ?",
                    (insc["passaro_id"],)
                ).fetchone()
            
            if not passaro:
                flash("Pássaro não encontrado.", "erro")
                return redirect(request.referrer or url_for("area_socio"))
        
        db.cancelar_inscricao(inscricao_id, passaro["socio_id"])
        flash("Inscrição cancelada pelo administrador com sucesso!", "sucesso")
    except (regras.ErroValidacao, ValueError) as e:
        flash(str(e), "erro")
    return redirect(request.referrer or url_for("area_socio"))


# ================================================================
# ÁREA DO SÓCIO
# ================================================================
@app.route("/area-socio")
@login_obrigatorio
def area_socio():
    socio_id = session["socio_id"]
    socio = db.obter_socio(socio_id)
    passaros = db.listar_passaros_do_socio(socio_id)
    for p in passaros:
        try:
            p["categoria"] = regras.calcular_categoria(p["codigo_ave"])
        except regras.ErroValidacao:
            p["categoria"] = None
    
    inscricoes_agrupadas = db.listar_inscricoes_do_socio_por_categoria(socio_id)
    
    # Ordena as inscrições dentro de cada categoria por ordem (crescente)
    for categoria, inscricoes in inscricoes_agrupadas.items():
        inscricoes_agrupadas[categoria] = sorted(
            inscricoes,
            key=lambda x: (x["ordem"] is None, x["ordem"] if x["ordem"] is not None else float('inf'))
        )
    
    torneios = db.listar_torneios()
    festivos = db.listar_festivos()
    
    return render_template("area_socio.html", 
                          socio=socio,
                          passaros=passaros, 
                          inscricoes_agrupadas=inscricoes_agrupadas,
                          torneios=torneios, 
                          festivos=festivos)


@app.route("/area-socio/passaros/novo", methods=["GET", "POST"])
@login_obrigatorio
def novo_passaro():
    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        sigla_criador = request.form.get("sigla_criador", "").strip().upper()
        numero_anilha = request.form.get("numero_anilha", "").strip()
        ano_anilha = request.form.get("ano_anilha", "").strip()

        try:
            db.cadastrar_passaro(
                socio_id=session["socio_id"],
                nome=nome,
                sigla_criador=sigla_criador,
                numero_anilha=numero_anilha,
                ano_anilha=ano_anilha
            )
            return render_template("novo_passaro.html", 
                                 passaro_cadastrado=nome,
                                 nome="",
                                 sigla_criador="",
                                 numero_anilha="",
                                 ano_anilha="")
        except (regras.ErroValidacao, ValueError) as e:
            flash(str(e), "erro")
            return render_template("novo_passaro.html", 
                                 nome=nome,
                                 sigla_criador=sigla_criador,
                                 numero_anilha=numero_anilha,
                                 ano_anilha=ano_anilha)

    return render_template("novo_passaro.html")


@app.route("/area-socio/passaros/<int:passaro_id>/editar", methods=["GET", "POST"])
@login_obrigatorio
def editar_passaro(passaro_id):
    socio_id = session["socio_id"]
    passaro = db.obter_passaro(passaro_id)
    
    if passaro is None or passaro["socio_id"] != socio_id:
        flash("Pássaro não encontrado.", "erro")
        return redirect(url_for("area_socio"))
    
    with db._conexao() as conn:
        if db.usar_postgres:
            cur = conn.cursor()
            cur.execute("SELECT id FROM inscricoes WHERE passaro_id = %s LIMIT 1", (passaro_id,))
            participou = cur.fetchone()
        else:
            participou = conn.execute(
                "SELECT id FROM inscricoes WHERE passaro_id = ? LIMIT 1",
                (passaro_id,)
            ).fetchone()
    
    if participou:
        if request.method == "POST":
            nome = request.form.get("nome", "").strip()
            sigla_criador = request.form.get("sigla_criador", "").strip().upper()
            numero_anilha = request.form.get("numero_anilha", "").strip()
            ano_anilha = request.form.get("ano_anilha", "").strip()
            
            try:
                db.editar_passaro(socio_id, passaro_id, nome, sigla_criador, 
                                 numero_anilha, ano_anilha)
                flash("Solicitação de edição enviada para aprovação do administrador.", "sucesso")
                return redirect(url_for("area_socio"))
            except (regras.ErroValidacao, ValueError) as e:
                flash(str(e), "erro")
                return render_template("editar_passaro.html", passaro=passaro, 
                                      nome=nome, sigla_criador=sigla_criador,
                                      numero_anilha=numero_anilha, ano_anilha=ano_anilha)
        
        return render_template("editar_passaro.html", passaro=passaro, precisa_aprovacao=True)
    
    else:
        if request.method == "POST":
            nome = request.form.get("nome", "").strip()
            sigla_criador = request.form.get("sigla_criador", "").strip().upper()
            numero_anilha = request.form.get("numero_anilha", "").strip()
            ano_anilha = request.form.get("ano_anilha", "").strip()
            
            try:
                with db._conexao() as conn:
                    if db.usar_postgres:
                        cur = conn.cursor()
                        cur.execute("""
                            UPDATE passaros SET 
                                nome = %s, sigla_criador = %s, numero_anilha = %s,
                                ano_anilha = %s
                            WHERE id = %s AND socio_id = %s
                        """, (nome, sigla_criador, numero_anilha, ano_anilha, 
                              passaro_id, socio_id))
                    else:
                        conn.execute("""
                            UPDATE passaros SET 
                                nome = ?, sigla_criador = ?, numero_anilha = ?,
                                ano_anilha = ?
                            WHERE id = ? AND socio_id = ?
                        """, (nome, sigla_criador, numero_anilha, ano_anilha, 
                              passaro_id, socio_id))
                flash("Pássaro atualizado com sucesso!", "sucesso")
                return redirect(url_for("area_socio"))
            except (regras.ErroValidacao, ValueError) as e:
                flash(str(e), "erro")
                return render_template("editar_passaro.html", passaro=passaro,
                                      nome=nome, sigla_criador=sigla_criador,
                                      numero_anilha=numero_anilha, ano_anilha=ano_anilha)
        
        return render_template("editar_passaro.html", passaro=passaro, precisa_aprovacao=False)


@app.route("/area-socio/editar-perfil", methods=["GET", "POST"])
@login_obrigatorio
def editar_perfil():
    socio_id = session["socio_id"]
    socio = db.obter_socio(socio_id)
    
    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        nascimento = request.form.get("nascimento", "")
        sexo = request.form.get("sexo", "")
        criatorio = request.form.get("criatorio", "").strip()
        cep = request.form.get("cep", "").strip()
        endereco = request.form.get("endereco", "").strip()
        numero = request.form.get("numero", "").strip()
        complemento = request.form.get("complemento", "").strip()
        bairro = request.form.get("bairro", "").strip()
        cidade = request.form.get("cidade", "").strip()
        uf = request.form.get("uf", "").strip().upper()
        pais = request.form.get("pais", "").strip()
        ddi = request.form.get("ddi", "").strip()
        ddd = request.form.get("ddd", "").strip()
        celular = request.form.get("celular", "").strip()
        whatsapp = request.form.get("whatsapp", "").strip()
        email = request.form.get("email", "").strip()
        facebook = request.form.get("facebook", "").strip()
        instagram = request.form.get("instagram", "").strip()
        youtube = request.form.get("youtube", "").strip()
        exibir_dados = request.form.get("exibir_dados", "0") == "1"
        
        try:
            with db._conexao() as conn:
                if db.usar_postgres:
                    cur = conn.cursor()
                    cur.execute("""
                        UPDATE socios SET 
                            nome = %s, nascimento = %s, sexo = %s, criatorio = %s,
                            cep = %s, endereco = %s, numero = %s, complemento = %s,
                            bairro = %s, cidade = %s, uf = %s, pais = %s,
                            ddi = %s, ddd = %s, celular = %s, whatsapp = %s,
                            email = %s, facebook = %s, instagram = %s, youtube = %s,
                            exibir_dados = %s
                        WHERE id = %s
                    """, (
                        nome, nascimento, sexo, criatorio,
                        cep, endereco, numero, complemento,
                        bairro, cidade, uf, pais,
                        ddi, ddd, celular, whatsapp,
                        email, facebook, instagram, youtube,
                        1 if exibir_dados else 0,
                        socio_id
                    ))
                else:
                    conn.execute("""
                        UPDATE socios SET 
                            nome = ?, nascimento = ?, sexo = ?, criatorio = ?,
                            cep = ?, endereco = ?, numero = ?, complemento = ?,
                            bairro = ?, cidade = ?, uf = ?, pais = ?,
                            ddi = ?, ddd = ?, celular = ?, whatsapp = ?,
                            email = ?, facebook = ?, instagram = ?, youtube = ?,
                            exibir_dados = ?
                        WHERE id = ?
                    """, (
                        nome, nascimento, sexo, criatorio,
                        cep, endereco, numero, complemento,
                        bairro, cidade, uf, pais,
                        ddi, ddd, celular, whatsapp,
                        email, facebook, instagram, youtube,
                        1 if exibir_dados else 0,
                        socio_id
                    ))
            
            session["socio_nome"] = nome
            flash("Perfil atualizado com sucesso!", "sucesso")
            return redirect(url_for("area_socio"))
        except Exception as e:
            flash(f"Erro ao atualizar: {str(e)}", "erro")
    
    return render_template("editar_perfil.html", socio=socio)


@app.route("/area-socio/transferir", methods=["GET", "POST"])
@login_obrigatorio
def transferir_passaro():
    socio_id = session["socio_id"]
    
    if request.method == "POST":
        passaro_id = request.form.get("passaro_id", type=int)
        cpf_destino = request.form.get("cpf_destino", "").strip()
        
        try:
            db.solicitar_transferencia(passaro_id, socio_id, cpf_destino)
            flash("Solicitação de transferência enviada com sucesso!", "sucesso")
            return redirect(url_for("transferir_passaro"))
        except (regras.ErroValidacao, ValueError) as e:
            flash(str(e), "erro")
    
    passaros = db.listar_passaros_do_socio(socio_id)
    transferencias_pendentes = db.listar_transferencias_pendentes(socio_id)
    transferencias_enviadas = db.listar_transferencias_enviadas(socio_id)
    
    return render_template("transferir_passaro.html", 
                          passaros=passaros,
                          transferencias_pendentes=transferencias_pendentes,
                          transferencias_enviadas=transferencias_enviadas)


@app.route("/area-socio/transferencia/<int:transferencia_id>/aceitar", methods=["POST"])
@login_obrigatorio
def aceitar_transferencia(transferencia_id):
    socio_id = session["socio_id"]
    try:
        db.aceitar_transferencia(transferencia_id, socio_id)
        flash("Transferência aceita com sucesso! O pássaro agora é seu.", "sucesso")
    except (regras.ErroValidacao, ValueError) as e:
        flash(str(e), "erro")
    return redirect(url_for("transferir_passaro"))


@app.route("/area-socio/transferencia/<int:transferencia_id>/recusar", methods=["POST"])
@login_obrigatorio
def recusar_transferencia(transferencia_id):
    socio_id = session["socio_id"]
    try:
        db.recusar_transferencia(transferencia_id, socio_id)
        flash("Transferência recusada.", "sucesso")
    except (regras.ErroValidacao, ValueError) as e:
        flash(str(e), "erro")
    return redirect(url_for("transferir_passaro"))


@app.route("/area-socio/etapa/<int:etapa_id>", methods=["GET", "POST"])
@login_obrigatorio
def inscrever_na_etapa(etapa_id):
    socio_id = session["socio_id"]
    etapa = db.obter_etapa(etapa_id)
    if etapa is None:
        flash("Etapa não encontrada.", "erro")
        return redirect(url_for("area_socio"))

    torneio = db.obter_torneio(etapa["torneio_id"]) if etapa["torneio_id"] else None
    socio = db.obter_socio(socio_id)

    if not etapa["inscricoes_abertas"] and not session.get("admin_id"):
        flash("Inscrições para esta etapa estão encerradas.", "erro")
        return redirect(url_for("area_socio"))

    total_inscritos = db.contar_inscricoes_na_etapa(etapa_id)
    if etapa.get("limite_inscricoes") is not None:
        if total_inscritos >= etapa["limite_inscricoes"]:
            flash(f"Limite de {etapa['limite_inscricoes']} inscrições atingido.", "erro")
            return redirect(url_for("area_socio"))

    if request.method == "POST":
        passaros_ids = request.form.getlist("passaros_ids")
        if not passaros_ids:
            flash("Selecione pelo menos um pássaro.", "erro")
            return redirect(url_for("inscrever_na_etapa", etapa_id=etapa_id))
        
        try:
            passaros_ids = [int(pid) for pid in passaros_ids]
            inscritos, erros = db.inscrever_multiplos_passaros(etapa_id, passaros_ids, socio_id)
            
            if inscritos:
                flash(f"{len(inscritos)} pássaro(s) inscrito(s) com sucesso! Aguarde a confirmação do pagamento.", "sucesso")
            if erros:
                for erro in erros:
                    flash(erro, "erro")
        except (regras.ErroValidacao, ValueError) as e:
            flash(str(e), "erro")
        return redirect(url_for("area_socio"))

    # ================================================================
    # LISTA DE PÁSSAROS COMPATÍVEIS - CORRIGIDO PARA MISTO
    # ================================================================
    passaros = db.listar_passaros_do_socio(socio_id)
    compativeis = []
    for p in passaros:
        try:
            categoria_passaro = regras.calcular_categoria(p["codigo_ave"])
            categoria_etapa = etapa["categoria"]
            
            # Se a etapa for MISTO, aceita QUALQUER pássaro (filhote OU adulto)
            if categoria_etapa == "MISTO":
                compativeis.append(p)
            # Senão, verifica se a categoria bate exatamente
            elif categoria_passaro == categoria_etapa:
                compativeis.append(p)
        except regras.ErroValidacao:
            pass

    inscritos = db.listar_inscritos_na_etapa(etapa_id)
    ja_inscritos_ids = {insc["passaro_id"] for insc in inscritos}
    ja_inscritos_cpf = db.contar_inscricoes_do_cpf_na_etapa(etapa_id, socio["cpf"])

    return render_template(
        "inscrever_etapa.html",
        etapa=etapa,
        torneio=torneio,
        passaros=compativeis,
        inscritos=inscritos,
        ja_inscritos_ids=ja_inscritos_ids,
        ja_inscritos_cpf=ja_inscritos_cpf,
        limite_por_cpf=etapa["limite_por_cpf"],
        total_inscritos=total_inscritos
    )


# ================================================================
# INSCRIÇÕES - CANCELAR
# ================================================================
@app.route("/area-socio/inscricao/<int:inscricao_id>/cancelar", methods=["POST"])
@login_obrigatorio
def cancelar_inscricao(inscricao_id):
    socio_id = session["socio_id"]
    try:
        db.cancelar_inscricao(inscricao_id, socio_id)
        flash("Inscrição cancelada com sucesso!", "sucesso")
    except (regras.ErroValidacao, ValueError) as e:
        flash(str(e), "erro")
    return redirect(url_for("area_socio"))


@app.route("/area-socio/pagamento/<int:inscricao_id>", methods=["GET", "POST"])
@login_obrigatorio
def enviar_comprovante(inscricao_id):
    socio_id = session["socio_id"]
    
    inscricoes = db.listar_inscricoes_do_socio(socio_id)
    inscricao = None
    for ins in inscricoes:
        if ins["inscricao_id"] == inscricao_id:
            inscricao = ins
            break
    
    if not inscricao:
        flash("Inscrição não encontrada.", "erro")
        return redirect(url_for("area_socio"))
    
    if request.method == "POST":
        comprovante = request.form.get("comprovante", "").strip()
        if not comprovante:
            flash("Por favor, informe o comprovante.", "erro")
            return render_template("enviar_comprovante.html", inscricao=inscricao)
        
        try:
            db.registrar_pagamento(inscricao_id, comprovante)
            flash("Comprovante enviado com sucesso! Aguarde a confirmação do administrador.", "sucesso")
            return redirect(url_for("area_socio"))
        except ValueError as e:
            flash(str(e), "erro")
    
    return render_template("enviar_comprovante.html", inscricao=inscricao)


# ================================================================
# API - BUSCA CEP
# ================================================================
@app.route("/api/cep/<cep>")
def buscar_cep(cep):
    try:
        import requests
        response = requests.get(f"https://viacep.com.br/ws/{cep}/json/")
        if response.status_code == 200:
            return jsonify(response.json())
        return jsonify({"erro": "CEP não encontrado"}), 404
    except:
        return jsonify({"erro": "Erro ao buscar CEP"}), 500


# ================================================================
# VERIFICAR PAGAMENTOS PENDENTES
# ================================================================
def verificar_pagamentos():
    try:
        logger.info("🔍 Verificando pagamentos pendentes...")
        cancelados = db.verificar_pagamentos_pendentes()
        if cancelados > 0:
            logger.info(f"✅ {cancelados} inscrições canceladas por falta de pagamento.")
        else:
            logger.info("✅ Nenhum pagamento pendente.")
    except Exception as e:
        logger.error(f"❌ Erro ao verificar pagamentos: {e}", exc_info=True)


# ================================================================
# INICIALIZAÇÃO
# ================================================================
if __name__ == "__main__":
    logger.info("🚀 Iniciando servidor Flask...")
    
    def verificar_periodicamente():
        while True:
            try:
                verificar_pagamentos()
            except Exception as e:
                logger.error(f"❌ Erro na thread: {e}", exc_info=True)
            time.sleep(3600)  # 1 hora
    
    thread = threading.Thread(target=verificar_periodicamente, daemon=True)
    thread.start()
    logger.info("✅ Thread de verificação iniciada")
    
    app.run(host="0.0.0.0", port=5000, debug=False)
